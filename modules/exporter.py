from __future__ import annotations

from copy import copy
from io import BytesIO
from typing import Optional

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SHEET_ORDER = [
    "账户总览",
    "AI 详情报告",
    "动作建议清单",
    "否定词清单",
    "暂停清单",
    "调价清单",
    "精准投放机会",
    "账户总数据明细",
    "优先级清单",
]


def build_excel_report(
    overview: pd.DataFrame,
    ai_report: pd.DataFrame,
    actions: pd.DataFrame,
    negative_keywords: pd.DataFrame,
    pause_list: pd.DataFrame,
    bid_adjustments: pd.DataFrame,
    exact_opportunities: pd.DataFrame,
    account_data: pd.DataFrame,
    priority_list: pd.DataFrame,
    aggregations: Optional[dict[str, pd.DataFrame]] = None,
) -> bytes:
    output = BytesIO()
    aggregations = aggregations or {}

    sheet_frames = {
        "账户总览": overview,
        "AI 详情报告": ai_report,
        "动作建议清单": actions,
        "否定词清单": negative_keywords,
        "暂停清单": pause_list,
        "调价清单": bid_adjustments,
        "精准投放机会": exact_opportunities,
        "账户总数据明细": account_data,
        "优先级清单": priority_list,
    }
    for dimension_name, dataframe in aggregations.items():
        sheet_frames[_safe_sheet_name(f"{dimension_name}聚合")] = dataframe

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, dataframe in sheet_frames.items():
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = writer.book
        for sheet_name, dataframe in sheet_frames.items():
            worksheet = workbook[sheet_name]
            _format_worksheet(worksheet, dataframe)
            _apply_tab_color(worksheet)

    return output.getvalue()


def _format_worksheet(worksheet, dataframe: pd.DataFrame) -> None:
    header_fill = PatternFill("solid", fgColor="22313F")
    header_font = Font(color="F7FAFC", bold=True)
    high_priority_fill = PatternFill("solid", fgColor="FCE4D6")
    medium_priority_fill = PatternFill("solid", fgColor="FFF2CC")
    low_priority_fill = PatternFill("solid", fgColor="E2F0D9")
    high_priority_font = Font(name="Arial", color="9C0006", bold=True)
    thin_border = Border(bottom=Side(style="thin", color="D9E2EC"))

    worksheet.freeze_panes = "A2"
    if dataframe.shape[0] >= 1 and dataframe.shape[1] >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    for row in worksheet.iter_rows():
        for cell in row:
            cell.font = copy(cell.font)
            cell.font = Font(
                name="Arial",
                size=cell.font.sz or 11,
                bold=cell.font.bold,
                italic=cell.font.italic,
                color=cell.font.color,
            )
            cell.border = thin_border
            alignment = copy(cell.alignment)
            alignment.vertical = "top"
            cell.alignment = alignment

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if worksheet.title == "AI 详情报告":
        worksheet.column_dimensions["A"].width = 24
        worksheet.column_dimensions["B"].width = 120
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                alignment = copy(cell.alignment)
                alignment.wrap_text = True
                alignment.vertical = "top"
                cell.alignment = alignment
        return

    for index, column_name in enumerate(dataframe.columns, start=1):
        letter = get_column_letter(index)
        sample_values = dataframe[column_name].astype(str).head(100).tolist() if not dataframe.empty else []
        max_length = max([len(str(column_name)), *[len(value) for value in sample_values]], default=12)
        worksheet.column_dimensions[letter].width = min(max(max_length + 4, 12), 42)

        number_format = _number_format_for(column_name)
        if number_format:
            for cell in worksheet[letter][1:]:
                cell.number_format = number_format
                if column_name in {"CTR", "CVR", "ACOS", "ROAS", "Spend", "Sales", "CPC", "Budget", "优先级评分"}:
                    alignment = copy(cell.alignment)
                    alignment.horizontal = "right"
                    cell.alignment = alignment

        if column_name in {"原因", "Reason", "报告内容", "诊断规则"}:
            worksheet.column_dimensions[letter].width = min(max(worksheet.column_dimensions[letter].width, 28), 80)
            for cell in worksheet[letter][1:]:
                alignment = copy(cell.alignment)
                alignment.wrap_text = True
                cell.alignment = alignment

    _highlight_priority_rows(
        worksheet,
        dataframe,
        high_priority_fill,
        medium_priority_fill,
        low_priority_fill,
        high_priority_font,
    )
    _add_score_color_scale(worksheet, dataframe)


def _highlight_priority_rows(
    worksheet,
    dataframe: pd.DataFrame,
    high_fill: PatternFill,
    medium_fill: PatternFill,
    low_fill: PatternFill,
    font: Font,
) -> None:
    if "优先级" not in dataframe.columns or dataframe.empty:
        return

    priority_column_index = list(dataframe.columns).index("优先级") + 1
    for row_index in range(2, len(dataframe) + 2):
        priority_value = worksheet.cell(row=row_index, column=priority_column_index).value
        if priority_value not in {"高", "中", "低"}:
            continue
        fill = {"高": high_fill, "中": medium_fill, "低": low_fill}[priority_value]
        for cell in worksheet[row_index]:
            cell.fill = fill
        if priority_value == "高":
            worksheet.cell(row=row_index, column=priority_column_index).font = font


def _add_score_color_scale(worksheet, dataframe: pd.DataFrame) -> None:
    if dataframe.empty or "优先级评分" not in dataframe.columns:
        return
    column_index = list(dataframe.columns).index("优先级评分") + 1
    letter = get_column_letter(column_index)
    cell_range = f"{letter}2:{letter}{len(dataframe) + 1}"
    worksheet.conditional_formatting.add(
        cell_range,
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="E2F0D9",
            mid_type="num",
            mid_value=60,
            mid_color="FFF2CC",
            end_type="num",
            end_value=100,
            end_color="F4B183",
        ),
    )


def _number_format_for(column_name: str) -> Optional[str]:
    if column_name in {"CTR", "CVR", "ACOS"}:
        return "0.00%"
    if column_name in {"Spend", "Sales", "CPC", "数值", "Budget", "总花费", "总销售额"}:
        return '"$"#,##0.00'
    if column_name == "ROAS":
        return "0.00"
    if column_name in {"Impressions", "Clicks", "Orders", "总曝光", "总点击", "总订单"}:
        return "#,##0"
    if column_name == "优先级评分":
        return "0"
    return None


def _safe_sheet_name(sheet_name: str) -> str:
    return sheet_name.replace("/", "-").replace("\\", "-")[:31]


def _apply_tab_color(worksheet) -> None:
    colors = {
        "账户总览": "22313F",
        "AI 详情报告": "5B9BD5",
        "动作建议清单": "ED7D31",
        "否定词清单": "C00000",
        "暂停清单": "7030A0",
        "调价清单": "70AD47",
        "精准投放机会": "00A6A6",
        "账户总数据明细": "7F7F7F",
        "优先级清单": "FFC000",
    }
    worksheet.sheet_properties.tabColor = colors.get(worksheet.title, "9EADBA")
